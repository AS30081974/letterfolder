import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import PyPDF2
from openpyxl import Workbook, load_workbook
import glob
import os
import threading
import subprocess
import time
import winreg
import ctypes
from ctypes import wintypes
from pathlib import Path


class PDFAddressExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("� Address Extractor Pro")
        self.root.geometry("900x700")
        self.root.configure(bg='#f8f9fa')
        self.root.resizable(True, True)
        
        # Set window icon and styling
        try:
            # Try to set a custom icon if available
            icon_path = os.path.join(os.path.dirname(__file__), "app_icon.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
            else:
                # Create a simple text-based icon using tkinter's built-in method
                self.root.iconname("AddressExtractor")
        except:
            pass
        
        # Variables
        self.pdf_folder_path = tk.StringVar()
        self.excel_file_path = tk.StringVar(value="addresses.xlsx")
        self.extraction_running = False
        self.printing_running = False
        self.adobe_path = None
        
        self.setup_modern_styles()
        self.setup_ui()
        self.find_adobe_reader()
        
    def setup_modern_styles(self):
        """Configure modern ttk styles"""
        self.style = ttk.Style()
        
        # Use a modern theme as base
        self.style.theme_use('clam')
        
        # Configure modern colors
        colors = {
            'primary': '#2563eb',      # Blue
            'primary_dark': '#1d4ed8',
            'secondary': '#64748b',    # Slate
            'success': '#059669',      # Green
            'danger': '#dc2626',       # Red
            'warning': '#d97706',      # Orange
            'light': '#f8fafc',        # Light gray
            'dark': '#1e293b',         # Dark gray
            'white': '#ffffff',
            'border': '#e2e8f0'
        }
        
        # Configure button styles
        self.style.configure('Primary.TButton',
                           background=colors['primary'],
                           foreground='white',
                           borderwidth=0,
                           focuscolor='none',
                           padding=(20, 12))
        
        self.style.map('Primary.TButton',
                      background=[('active', colors['primary_dark']),
                                ('pressed', colors['primary_dark'])])
        
        self.style.configure('Secondary.TButton',
                           background=colors['secondary'],
                           foreground='white',
                           borderwidth=0,
                           focuscolor='none',
                           padding=(16, 10))
        
        self.style.map('Secondary.TButton',
                      background=[('active', colors['dark']),
                                ('pressed', colors['dark'])])
        
        self.style.configure('Success.TButton',
                           background=colors['success'],
                           foreground='white',
                           borderwidth=0,
                           focuscolor='none',
                           padding=(16, 10))
        
        self.style.map('Success.TButton',
                      background=[('active', '#047857'),
                                ('pressed', '#047857')])
        
        # Configure frame styles
        self.style.configure('Card.TFrame',
                           background='white',
                           relief='flat',
                           borderwidth=1)
        
        self.style.configure('Header.TFrame',
                           background=colors['primary'],
                           relief='flat')
        
        # Configure label styles
        self.style.configure('Title.TLabel',
                           background='white',
                           foreground=colors['dark'],
                           font=('Segoe UI', 24, 'bold'))
        
        self.style.configure('Subtitle.TLabel',
                           background='white',
                           foreground=colors['secondary'],
                           font=('Segoe UI', 11))
        
        self.style.configure('CardTitle.TLabel',
                           background='white',
                           foreground=colors['dark'],
                           font=('Segoe UI', 12, 'bold'))
        
        self.style.configure('FieldLabel.TLabel',
                           background='white',
                           foreground=colors['dark'],
                           font=('Segoe UI', 10))
        
        # Configure entry styles
        self.style.configure('Modern.TEntry',
                           fieldbackground='white',
                           borderwidth=2,
                           relief='solid',
                           insertcolor=colors['primary'],
                           padding=(12, 8))
        
        # Configure progressbar styles
        self.style.configure('Modern.Horizontal.TProgressbar',
                           background=colors['primary'],
                           troughcolor='#e5e7eb',
                           borderwidth=0,
                           lightcolor=colors['primary'],
                           darkcolor=colors['primary'])
        
        # Configure labelframe styles
        self.style.configure('Modern.TLabelframe',
                           background='white',
                           relief='flat',
                           borderwidth=0)
        
        self.style.configure('Modern.TLabelframe.Label',
                           background='white',
                           foreground=colors['dark'],
                           font=('Segoe UI', 12, 'bold'))
        
    def setup_ui(self):
        # Create main container with padding
        main_container = tk.Frame(self.root, bg='#f8f9fa')
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Configure grid weights
        main_container.grid_rowconfigure(1, weight=1)
        main_container.grid_columnconfigure(0, weight=1)
        
        # Header section
        self.create_header(main_container)
        
        # Main content area
        content_frame = tk.Frame(main_container, bg='#f8f9fa')
        content_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(20, 0))
        content_frame.grid_rowconfigure(2, weight=1)
        content_frame.grid_columnconfigure(0, weight=1)
        
        # File selection card
        self.create_file_selection_card(content_frame)
        
        # Action buttons card
        self.create_action_buttons_card(content_frame)
        
        # Status and progress card
        self.create_status_card(content_frame)
        
        # Set default folder
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.pdf_folder_path.set(current_dir)
        
        self.log_message("🚀 Address Extractor Pro started. Ready to extract addresses from PDFs.")
        
    def create_header(self, parent):
        """Create modern header section"""
        header_frame = ttk.Frame(parent, style='Card.TFrame')
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        header_frame.grid_columnconfigure(0, weight=1)
        
        # Add padding inside the card
        header_content = tk.Frame(header_frame, bg='white')
        header_content.pack(fill=tk.BOTH, expand=True, padx=30, pady=25)
        
        # Title with icon
        title_frame = tk.Frame(header_content, bg='white')
        title_frame.pack(anchor=tk.W)
        
        title_label = ttk.Label(title_frame, text="🏢📋 Address Extractor Pro", style='Title.TLabel')
        title_label.pack(side=tk.LEFT)
        
        # Subtitle
        subtitle_label = ttk.Label(header_content, 
                                 text="Extract addresses from PDF files and save to Excel spreadsheets", 
                                 style='Subtitle.TLabel')
        subtitle_label.pack(anchor=tk.W, pady=(5, 0))
        
    def create_file_selection_card(self, parent):
        """Create file selection card"""
        card_frame = ttk.Frame(parent, style='Card.TFrame')
        card_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        card_frame.grid_columnconfigure(1, weight=1)
        
        # Card content with padding
        content = tk.Frame(card_frame, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)
        content.grid_columnconfigure(1, weight=1)
        
        # Card title
        ttk.Label(content, text="📁 File Selection", style='CardTitle.TLabel').grid(
            row=0, column=0, columnspan=3, sticky=tk.W, pady=(0, 20))
        
        # PDF Folder Selection
        ttk.Label(content, text="PDF Folder:", style='FieldLabel.TLabel').grid(
            row=1, column=0, sticky=tk.W, pady=(0, 15), padx=(0, 15))
        
        pdf_entry = ttk.Entry(content, textvariable=self.pdf_folder_path, 
                             style='Modern.TEntry', font=('Segoe UI', 10))
        pdf_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(0, 15), padx=(0, 10))
        
        pdf_browse_btn = ttk.Button(content, text="📂 Browse", 
                                   command=self.browse_pdf_folder, style='Secondary.TButton')
        pdf_browse_btn.grid(row=1, column=2, pady=(0, 15))
        
        # Excel File Selection
        ttk.Label(content, text="Excel File:", style='FieldLabel.TLabel').grid(
            row=2, column=0, sticky=tk.W, padx=(0, 15))
        
        excel_entry = ttk.Entry(content, textvariable=self.excel_file_path, 
                               style='Modern.TEntry', font=('Segoe UI', 10))
        excel_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        excel_browse_btn = ttk.Button(content, text="📊 Browse", 
                                     command=self.browse_excel_file, style='Secondary.TButton')
        excel_browse_btn.grid(row=2, column=2)
        
    def create_action_buttons_card(self, parent):
        """Create action buttons card"""
        card_frame = ttk.Frame(parent, style='Card.TFrame')
        card_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Card content with padding
        content = tk.Frame(card_frame, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)
        
        # Card title
        ttk.Label(content, text="⚡ Actions", style='CardTitle.TLabel').pack(
            anchor=tk.W, pady=(0, 20))
        
        # Buttons container
        buttons_frame = tk.Frame(content, bg='white')
        buttons_frame.pack(anchor=tk.W)
        
        # Extract button with icon
        self.extract_button = ttk.Button(buttons_frame, text="🔍 Extract Addresses", 
                                        command=self.extract_addresses_threaded,
                                        style='Primary.TButton')
        self.extract_button.pack(side=tk.LEFT, padx=(0, 15))
        
        # Print buttons frame
        print_frame = tk.Frame(buttons_frame, bg='white')
        print_frame.pack(side=tk.LEFT, padx=(0, 15))
        
        # Print Visible button
        self.print_visible_button = ttk.Button(print_frame, text="🖨️👁️ Print (Visible)", 
                                              command=self.print_pdfs_visible,
                                              style='Success.TButton')
        self.print_visible_button.pack(side=tk.TOP, pady=(0, 5))
        
        # Stop print button (initially hidden)
        self.stop_print_button = ttk.Button(buttons_frame, text="⏹️ Stop Printing", 
                                           command=self.stop_printing,
                                           style='Secondary.TButton')
        
        # Clear button
        self.clear_button = ttk.Button(buttons_frame, text="🧹 Clear Spreadsheet", 
                                      command=self.clear_spreadsheet,
                                      style='Secondary.TButton')
        self.clear_button.pack(side=tk.LEFT)
        
        # Progress bar (initially hidden)
        self.progress_frame = tk.Frame(content, bg='white')
        self.progress_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Label(self.progress_frame, text="Processing...", 
                 style='FieldLabel.TLabel').pack(anchor=tk.W, pady=(0, 5))
        
        self.progress = ttk.Progressbar(self.progress_frame, mode='indeterminate', 
                                       style='Modern.Horizontal.TProgressbar')
        self.progress.pack(fill=tk.X)
        
        # Hide progress initially
        self.progress_frame.pack_forget()
        
    def create_status_card(self, parent):
        """Create status and log card"""
        card_frame = ttk.Frame(parent, style='Card.TFrame')
        card_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        card_frame.grid_rowconfigure(1, weight=1)
        card_frame.grid_columnconfigure(0, weight=1)
        
        # Card content with padding
        content = tk.Frame(card_frame, bg='white')
        content.pack(fill=tk.BOTH, expand=True, padx=25, pady=20)
        content.grid_rowconfigure(1, weight=1)
        content.grid_columnconfigure(0, weight=1)
        
        # Card title
        ttk.Label(content, text="📋 Status Log", style='CardTitle.TLabel').grid(
            row=0, column=0, sticky=tk.W, pady=(0, 15))
        
        # Status text area with modern styling
        text_frame = tk.Frame(content, bg='#f8fafc', relief='solid', bd=1)
        text_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        text_frame.grid_rowconfigure(0, weight=1)
        text_frame.grid_columnconfigure(0, weight=1)
        
        self.status_text = scrolledtext.ScrolledText(
            text_frame, 
            height=12, 
            width=70,
            bg='#f8fafc',
            fg='#334155',
            font=('Consolas', 10),
            relief='flat',
            borderwidth=0,
            insertbackground='#2563eb',
            selectbackground='#dbeafe',
            wrap=tk.WORD
        )
        self.status_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=15, pady=15)
        
    def browse_pdf_folder(self):
        folder = filedialog.askdirectory(title="Select PDF Folder")
        if folder:
            self.pdf_folder_path.set(folder)
            self.log_message(f"PDF folder selected: {folder}")
    
    def browse_excel_file(self):
        file = filedialog.asksaveasfilename(
            title="Select Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file:
            self.excel_file_path.set(file)
            self.log_message(f"Excel file selected: {file}")
    
    def log_message(self, message):
        self.status_text.insert(tk.END, f"{message}\n")
        self.status_text.see(tk.END)
        self.root.update_idletasks()
    
    def extract_text_from_pdfs(self, folder_path):
        """Extract text from PDF files in the specified folder"""
        pdf_texts = {}
        
        # Find all PDF files in the folder
        pdf_pattern = os.path.join(folder_path, "*.pdf")
        pdf_files = glob.glob(pdf_pattern)
        
        if not pdf_files:
            self.log_message("No PDF files found in the selected folder.")
            return pdf_texts
        
        self.log_message(f"Found {len(pdf_files)} PDF files to process.")
        
        for file_path in pdf_files:
            try:
                self.log_message(f"Processing: {os.path.basename(file_path)}")
                
                with open(file_path, 'rb') as pdf_file:
                    reader = PyPDF2.PdfReader(pdf_file)
                    text = ''
                    
                    for page_num in range(len(reader.pages)):
                        page = reader.pages[page_num]
                        text += page.extract_text()
                    
                    # Extract text between ".com" and "Dear"
                    start = text.find("uk_team_gbmailgps@lilly.com")
                    end = text.find("Dear")
                    
                    if start != -1 and end != -1:
                        lines = text[start + 5:end].split('\n')
                        lines = [line.strip() for line in lines if line.strip() != ""]
                        extracted_text = '\n'.join(lines[2:]) if len(lines) > 2 else '\n'.join(lines)
                        pdf_texts[os.path.basename(file_path)] = extracted_text
                        self.log_message(f"Successfully extracted address from {os.path.basename(file_path)}")
                    else:
                        self.log_message(f"Could not find address markers in {os.path.basename(file_path)}")
                        
            except Exception as e:
                self.log_message(f"Error processing {os.path.basename(file_path)}: {str(e)}")
        
        return pdf_texts
    
    def extract_addresses_threaded(self):
        """Run extraction in a separate thread to prevent UI freezing"""
        if self.extraction_running:
            self.log_message("Extraction is already running. Please wait...")
            return
            
        thread = threading.Thread(target=self.extract_addresses)
        thread.daemon = True
        thread.start()
    
    def extract_addresses(self):
        """Main extraction function"""
        try:
            self.extraction_running = True
            self.extract_button.config(state='disabled')
            self.progress_frame.pack(fill=tk.X, pady=(20, 0))
            self.progress.start()
            
            # Validate inputs
            if not self.pdf_folder_path.get():
                messagebox.showerror("Error", "Please select a PDF folder.")
                return
            
            if not self.excel_file_path.get():
                messagebox.showerror("Error", "Please specify an Excel file.")
                return
            
            if not os.path.exists(self.pdf_folder_path.get()):
                messagebox.showerror("Error", "PDF folder does not exist.")
                return
            
            self.log_message("Starting address extraction...")
            
            # Extract text from PDFs
            pdf_texts = self.extract_text_from_pdfs(self.pdf_folder_path.get())
            
            if not pdf_texts:
                self.log_message("No addresses were extracted.")
                return
            
            # Handle Excel file
            excel_path = self.excel_file_path.get()
            
            # Create or load workbook
            if os.path.exists(excel_path):
                try:
                    wb = load_workbook(excel_path)
                    self.log_message(f"Loaded existing Excel file: {excel_path}")
                except Exception as e:
                    self.log_message(f"Error loading Excel file, creating new one: {str(e)}")
                    wb = Workbook()
            else:
                wb = Workbook()
                self.log_message(f"Created new Excel file: {excel_path}")
            
            sheet = wb.active
            if sheet.title == "Sheet":
                sheet.title = "Addresses"
            
            # Add headers if the sheet is empty
            if sheet.max_row == 1 and sheet['A1'].value is None:
                sheet['A1'] = "PDF File"
                sheet['B1'] = "Extracted Address"
                current_row = 2
            else:
                current_row = sheet.max_row + 1
            
            # Write extracted addresses to Excel
            self.log_message("Writing addresses to Excel...")
            
            for pdf_file, address in pdf_texts.items():
                sheet.cell(row=current_row, column=1).value = pdf_file
                sheet.cell(row=current_row, column=2).value = address
                current_row += 1
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(excel_path)
            wb.close()
            
            self.log_message(f"Successfully extracted {len(pdf_texts)} addresses and saved to {excel_path}")
            messagebox.showinfo("Success", f"Extracted {len(pdf_texts)} addresses successfully!")
            
        except Exception as e:
            error_msg = f"An error occurred: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("Error", error_msg)
        
        finally:
            self.extraction_running = False
            self.extract_button.config(state='normal')
            self.progress.stop()
            self.progress_frame.pack_forget()
    
    def clear_spreadsheet(self):
        """Clear the contents of the Excel spreadsheet"""
        try:
            excel_path = self.excel_file_path.get()
            
            if not excel_path:
                messagebox.showerror("Error", "Please specify an Excel file.")
                return
            
            # Confirm action
            result = messagebox.askyesno(
                "Confirm Clear", 
                f"Are you sure you want to clear all data from {os.path.basename(excel_path)}?\n\nThis action cannot be undone."
            )
            
            if not result:
                return
            
            # Create a new workbook or clear existing one
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Addresses"
            
            # Add headers
            sheet['A1'] = "PDF File"
            sheet['B1'] = "Extracted Address"
            
            # Save the workbook
            wb.save(excel_path)
            wb.close()
            
            self.log_message(f"Cleared spreadsheet: {excel_path}")
            messagebox.showinfo("Success", "Spreadsheet cleared successfully!")
            
        except Exception as e:
            error_msg = f"Error clearing spreadsheet: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("Error", error_msg)
    
    def find_adobe_reader(self):
        """Find Adobe Reader or Acrobat installation path"""
        try:
            # Common Adobe paths
            adobe_paths = [
                r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
                r"C:\Program Files (x86)\Adobe\Acrobat DC\Acrobat\Acrobat.exe",
                r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
                r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe",
                r"C:\Program Files\Adobe\Reader 11.0\Reader\AcroRd32.exe",
                r"C:\Program Files (x86)\Adobe\Reader 11.0\Reader\AcroRd32.exe"
            ]
            
            # Check each path
            for path in adobe_paths:
                if os.path.exists(path):
                    self.adobe_path = path
                    app_name = "Acrobat" if "Acrobat.exe" in path else "Adobe Reader"
                    self.log_message(f"✅ Found {app_name} for printing functionality")
                    return
            
            # Try to find through registry
            try:
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, 
                                   r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\AcroRd32.exe")
                self.adobe_path, _ = winreg.QueryValueEx(key, "")
                winreg.CloseKey(key)
                self.log_message(f"✅ Found Adobe Reader through registry for printing")
                return
            except:
                pass
            
            self.log_message("⚠️ Adobe Reader/Acrobat not found. Print functionality will use system default.")
            
        except Exception as e:
            self.log_message(f"❌ Error finding Adobe: {e}")
    
    def print_pdfs_visible(self):
        """Print PDFs with visible windows and automation"""
        self.print_pdfs_threaded()
    
    def print_pdfs_threaded(self):
        """Run PDF printing in a separate thread to prevent UI freezing"""
        if self.printing_running:
            self.log_message("PDF printing is already running. Please wait...")
            return
            
        if self.extraction_running:
            self.log_message("Please wait for address extraction to complete before printing.")
            return
            
        thread = threading.Thread(target=self.print_pdfs)
        thread.daemon = True
        thread.start()
    
    def print_pdfs(self):
        """Main PDF printing function"""
        try:
            self.printing_running = True
            self.print_visible_button.config(state='disabled')
            self.stop_print_button.pack(side=tk.LEFT, padx=(0, 15))
            self.progress_frame.pack(fill=tk.X, pady=(20, 0))
            self.progress.start()
            
            # Validate inputs
            if not self.pdf_folder_path.get():
                messagebox.showerror("Error", "Please select a PDF folder.")
                return
            
            if not os.path.exists(self.pdf_folder_path.get()):
                messagebox.showerror("Error", "PDF folder does not exist.")
                return
            
            # Find PDF files
            pdf_pattern = os.path.join(self.pdf_folder_path.get(), "*.pdf")
            pdf_files = glob.glob(pdf_pattern)
            
            if not pdf_files:
                messagebox.showinfo("Info", "No PDF files found in the selected folder.")
                return
            
            # Ask user for confirmation
            result = messagebox.askyesno(
                "Print PDFs - Visible Mode", 
                f"Ready to print {len(pdf_files)} PDF files using visible automation?\n\n"
                f"📌 Stapling Setup Instructions:\n"
                f"    • Open Printers & Scanners settings\n"
                f"    • Select XWCSmartPrint\n"
                f"    • Click 'Print Properties'\n"
                f"    • Under 'Presets', select '1 Staple, 2-Sided' and apply\n\n"
                f"Process Information:\n"
                f"    • Each PDF will open visibly on screen\n"
                f"    • Print commands will be automated (Ctrl+P, Enter)\n"
                f"    • Each file takes approximately 10-15 seconds\n\n"
                f"⚠️ IMPORTANT: Please do not use keyboard or mouse during printing.\n\n"
                f"Proceed with printing?"
            )
            
            if not result:
                return
            
            self.log_message(f"🖨️ Starting VISIBLE printing of {len(pdf_files)} PDF files...")
            self.log_message(f"⚠️ IMPORTANT: Please do not use your computer during printing!")
            
            printed_count = 0
            failed_count = 0
            
            for i, pdf_file in enumerate(sorted(pdf_files)):
                if not self.printing_running:  # Check if stopped
                    self.log_message("🛑 Printing stopped by user.")
                    break
                
                try:
                    filename = os.path.basename(pdf_file)
                    self.log_message(f"🖨️ [{i+1}/{len(pdf_files)}] Processing {filename}...")
                    
                    success = self.print_single_pdf_visible(pdf_file)
                    
                    if success:
                        printed_count += 1
                        self.log_message(f"   ✅ Successfully processed")
                    else:
                        failed_count += 1
                        self.log_message(f"   ❌ Failed to process")
                    
                    # Small delay between files
                    if i < len(pdf_files) - 1:
                        self.log_message(f"   ⏱️ Waiting 2 seconds before next file...")
                        for j in range(20):  # 2 seconds with responsiveness check
                            if not self.printing_running:
                                break
                            time.sleep(0.1)
                            self.root.update_idletasks()
                        
                except Exception as e:
                    failed_count += 1
                    self.log_message(f"   ❌ Error: {e}")
            
            if self.printing_running:  # Completed normally
                self.log_message(f"🎉 VISIBLE printing completed!")
                self.log_message(f"✅ Successfully processed: {printed_count} files")
                if failed_count > 0:
                    self.log_message(f"❌ Failed: {failed_count} files")
                
                messagebox.showinfo("Success", 
                    f"Visible printing completed!\n\n"
                    f"✅ Successfully processed: {printed_count} files\n"
                    f"❌ Failed: {failed_count} files\n\n"
                    f"Check your printer queue for print jobs.")
            
        except Exception as e:
            error_msg = f"❌ Printing error: {e}"
            self.log_message(error_msg)
            messagebox.showerror("Error", error_msg)
        
        finally:
            self.printing_running = False
            self.print_visible_button.config(state='normal')
            self.stop_print_button.pack_forget()
            self.progress.stop()
            self.progress_frame.pack_forget()
    
    def print_single_pdf_visible(self, pdf_path):
        """Print a single PDF file with visible automation"""
        try:
            if self.adobe_path and os.path.exists(self.adobe_path):
                # Open Adobe Reader/Acrobat normally (visible window)
                self.log_message(f"   📖 Opening PDF in Adobe...")
                
                cmd = [self.adobe_path, pdf_path]
                process = subprocess.Popen(cmd)
                
                # Wait for Adobe to load
                time.sleep(3)
                
                # Focus on Adobe window
                self.focus_adobe_window()
                time.sleep(0.5)
                
                # Send Ctrl+P to print
                self.log_message(f"   🖨️ Sending print command (Ctrl+P)...")
                self.send_ctrl_p()
                time.sleep(2)
                
                # Send Enter to confirm print
                self.log_message(f"   ✅ Confirming print (Enter)...")
                self.send_enter()
                time.sleep(2)
                
                # Close Adobe window
                self.log_message(f"   🔄 Closing Adobe...")
                self.close_adobe_window()
                
                return True
                
            else:
                # Fallback: Use system default PDF handler
                self.log_message(f"   📄 Opening PDF with system default...")
                os.startfile(pdf_path)
                time.sleep(3)
                
                # Try to send print command
                self.send_ctrl_p()
                time.sleep(2)
                self.send_enter()
                time.sleep(2)
                
                return True
                    
        except Exception as e:
            self.log_message(f"   ❌ Visible print error: {e}")
            return False
    
    def stop_printing(self):
        """Stop the printing process"""
        self.printing_running = False
        self.log_message("🛑 Stopping printing process...")
        
        # Close any open Adobe windows
        try:
            self.close_adobe_processes()
        except:
            pass
    
    def focus_adobe_window(self):
        """Focus on Adobe Reader/Acrobat window"""
        try:
            user32 = ctypes.windll.user32
            
            def enum_windows_callback(hwnd, lParam):
                if user32.IsWindowVisible(hwnd):
                    window_title = ctypes.create_unicode_buffer(512)
                    user32.GetWindowTextW(hwnd, window_title, 512)
                    title = window_title.value.lower()
                    
                    if any(app in title for app in ['adobe', 'acrobat', 'reader']):
                        user32.SetForegroundWindow(hwnd)
                        user32.ShowWindow(hwnd, 9)  # SW_RESTORE
                        return False  # Stop enumeration
                return True
            
            # Enumerate all windows
            WNDENUMPROC = ctypes.WINFUNCTYPE(ctypes.c_bool, wintypes.HWND, wintypes.LPARAM)
            user32.EnumWindows(WNDENUMPROC(enum_windows_callback), 0)
            
        except Exception as e:
            self.log_message(f"   ⚠️ Could not focus Adobe window: {e}")
    
    def send_ctrl_p(self):
        """Send Ctrl+P key combination"""
        try:
            user32 = ctypes.windll.user32
            VK_CONTROL = 0x11
            VK_P = 0x50
            
            user32.keybd_event(VK_CONTROL, 0, 0, 0)
            user32.keybd_event(VK_P, 0, 0, 0)
            user32.keybd_event(VK_P, 0, 2, 0)  # KEYEVENTF_KEYUP
            user32.keybd_event(VK_CONTROL, 0, 2, 0)
            
        except Exception as e:
            self.log_message(f"   ⚠️ Could not send Ctrl+P: {e}")
    
    def send_enter(self):
        """Send Enter key"""
        try:
            user32 = ctypes.windll.user32
            VK_RETURN = 0x0D
            
            user32.keybd_event(VK_RETURN, 0, 0, 0)
            user32.keybd_event(VK_RETURN, 0, 2, 0)
            
        except Exception as e:
            self.log_message(f"   ⚠️ Could not send Enter: {e}")
    
    def close_adobe_window(self):
        """Close Adobe window using Alt+F4"""
        try:
            user32 = ctypes.windll.user32
            VK_MENU = 0x12  # Alt key
            VK_F4 = 0x73
            
            user32.keybd_event(VK_MENU, 0, 0, 0)
            user32.keybd_event(VK_F4, 0, 0, 0)
            user32.keybd_event(VK_F4, 0, 2, 0)
            user32.keybd_event(VK_MENU, 0, 2, 0)
            
        except Exception as e:
            self.log_message(f"   ⚠️ Could not close Adobe window: {e}")
    
    def close_adobe_processes(self):
        """Force close Adobe processes"""
        try:
            subprocess.run(['taskkill', '/f', '/im', 'AcroRd32.exe'], 
                          capture_output=True, stderr=subprocess.DEVNULL)
            subprocess.run(['taskkill', '/f', '/im', 'Acrobat.exe'], 
                          capture_output=True, stderr=subprocess.DEVNULL)
        except Exception:
            pass


def main():
    root = tk.Tk()
    
    # Configure ttk styles for a modern look
    style = ttk.Style()
    style.theme_use('clam')  # Use a modern theme
    
    # Create custom style for accent button
    style.configure('Accent.TButton', foreground='white', background='#0078d4')
    style.map('Accent.TButton', 
             background=[('active', '#106ebe'), ('pressed', '#005a9e')])
    
    app = PDFAddressExtractorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
