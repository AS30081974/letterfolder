"""
Excel Processing Module
Handles Excel file operations for saving extracted PDF data.
"""

from openpyxl import Workbook, load_workbook
import os
from typing import Dict, List


class ExcelHandler:
    """Handles Excel file operations for saving extracted data."""
    
    def save_data_to_excel(self, pdf_texts: Dict[str, str], excel_path: str, log_callback=None) -> bool:
        """
        Save extracted PDF data to Excel file.
        
        Args:
            pdf_texts: Dictionary mapping PDF filenames to extracted text
            excel_path: Path to Excel file to save/update
            log_callback: Optional callback function for logging messages
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create or load workbook
            if os.path.exists(excel_path):
                try:
                    wb = load_workbook(excel_path)
                    if log_callback:
                        log_callback(f"Loaded existing Excel file: {excel_path}")
                except Exception as e:
                    if log_callback:
                        log_callback(f"Error loading Excel file, creating new one: {str(e)}")
                    wb = Workbook()
            else:
                wb = Workbook()
                if log_callback:
                    log_callback(f"Created new Excel file: {excel_path}")
            
            sheet = wb.active
            if sheet.title == "Sheet":
                sheet.title = "Extracted Data"
            
            # Add headers if the sheet is empty
            if sheet.max_row == 1 and sheet['A1'].value is None:
                sheet['A1'] = "PDF File"
                sheet['B1'] = "Extracted Data"
                current_row = 2
            else:
                current_row = sheet.max_row + 1
            
            # Write extracted data to Excel
            if log_callback:
                log_callback("Writing extracted data to Excel...")
            
            for pdf_file, data in pdf_texts.items():
                sheet.cell(row=current_row, column=1).value = pdf_file
                sheet.cell(row=current_row, column=2).value = data
                current_row += 1
            
            # Auto-adjust column widths
            self._adjust_column_widths(sheet)
            
            # Save the workbook
            wb.save(excel_path)
            wb.close()
            
            if log_callback:
                log_callback(f"Successfully saved {len(pdf_texts)} entries to {excel_path}")
            
            return True
            
        except Exception as e:
            if log_callback:
                log_callback(f"Error saving to Excel: {str(e)}")
            return False
    
    def clear_spreadsheet(self, excel_path: str, log_callback=None) -> bool:
        """
        Clear the contents of the Excel spreadsheet.
        
        Args:
            excel_path: Path to Excel file to clear
            log_callback: Optional callback function for logging messages
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create a new workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Extracted Data"
            
            # Add headers
            sheet['A1'] = "PDF File"
            sheet['B1'] = "Extracted Data"
            
            # Save the workbook
            wb.save(excel_path)
            wb.close()
            
            if log_callback:
                log_callback(f"Cleared spreadsheet: {excel_path}")
            
            return True
            
        except Exception as e:
            if log_callback:
                log_callback(f"Error clearing spreadsheet: {str(e)}")
            return False
    
    def append_addresses(self, addresses: List[str], excel_path: str) -> bool:
        """
        Append addresses to an existing Excel file.
        
        Args:
            addresses: List of address strings to append
            excel_path: Path to the existing Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Load the existing workbook
            wb = load_workbook(excel_path)
            sheet = wb.active
            
            # Find the next empty row
            next_row = sheet.max_row + 1
            
            # Add each address to a new row
            for address in addresses:
                sheet[f'A{next_row}'] = f"Appended_{next_row}"  # File identifier
                sheet[f'B{next_row}'] = address.strip()
                next_row += 1
            
            # Auto-size columns
            for column_cells in sheet.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                adjusted_width = min(max(length, 10), 100)
                column_letter = column_cells[0].column_letter
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(excel_path)
            wb.close()
            
            return True
            
        except Exception as e:
            print(f"Error appending to Excel file: {e}")
            return False
    
    def _adjust_column_widths(self, sheet):
        """Automatically adjust column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
